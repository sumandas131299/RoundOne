import os

from flask import Blueprint, json, jsonify,render_template, request
from google import genai
from google.genai import types
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook 

FILE_NAME = 'cx_reviews.xlsx'
load_dotenv() #loading the environment variables

feedbackBp = Blueprint('feedback',__name__)

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
client = genai.Client(api_key=GEMINI_API_KEY)

@feedbackBp.route('/feedback')
def feedbackIndex():
    return render_template('feedback.html')


@feedbackBp.route('/feedback/save', methods=['POST'])
def save_feedback():
    #audio = request.files.get('audio_file')
    i_type = request.form.get('type')
    diff = request.form.get('difficulty')
    question = request.form.get('question')
    print(i_type)
    print(diff)
    print(question)
    # 1. Save File Temporarily
    

    try:
        # 2. Upload to Gemini
# Change 'path' to 'file'
        #gemini_file = client.files.upload(file=file_path)

        # 3. Precise Schema for your HTML
        response_schema = {
            "type": "OBJECT",
            "properties": {
                "overall_score": {"type": "NUMBER"},
                "score_label": {"type": "STRING"},
                "transcript": {"type": "STRING"},
                "categories": {
                    "type": "ARRAY",
                    "items": {
                        "type": "OBJECT",
                        "properties": {
                            "title": {"type": "STRING"},
                            "score": {"type": "INTEGER"},
                            "percentage": {"type": "INTEGER"},
                            "status": {"type": "STRING"},
                            "feedback": {"type": "STRING"}
                        }
                    }
                }
            }
        }

        # prompt = f"""
        # Give a proper and appropriate answer for the interview question: '{question}'.
        # The interview type is {i_type} and difficulty is {diff}.
        # Evaluate based on: Answer Structure (STAR),Answer Clarity, Communication, and JD Match.
        # Return the feedback in the specified JSON format.
        # And if the audio is blank or transcript is not readable, then return response according to that like "your said nothing" and score 0.
        # overall score should be out of 10.
        # """

        prompt = f"""
        Question: "{question}" ({i_type}, {diff})

        TASK: Provide a appropriate and proper answer for this question. 
        OUTPUT: JSON ONLY score/10.
        """

        response = client.models.generate_content(
            model="gemini-2.5-flash-lite",
            contents=[ prompt],
            config=types.GenerateContentConfig(
                response_mime_type="application/json",
                response_schema=response_schema
            )
        )

        # 4. Clean up
        #os.remove(file_path)
        
        # 5. Return data to frontend
        feedback_data = json.loads(response.text)
        print(feedback_data)
        return jsonify({"status": "success", "data": feedback_data})

    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500





@feedbackBp.route('/feedback/save_review', methods=['POST'])
def save_review_excel():
    try:
        data = request.json
        career = data.get('career')
        rating = data.get('rating')
        comment = data.get('comment')

        # Check if file exists to determine if we need to create it or load it
        if os.path.exists(FILE_NAME):
            wb = load_workbook(FILE_NAME)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            # Add Headers for a new file
            ws.append(["Date", "Time","Career", "Rating", "Comment"])

        # Append the data row
        from datetime import datetime
        now = datetime.now()
        current_date = now.strftime("%Y-%m-%d")
        current_time = now.strftime("%H:%M:%S")
        ws.append([ current_date, current_time , career, rating, comment])
        
        wb.save(FILE_NAME)
        return jsonify({"status": "success", "message": "Feedback saved to Excel"})
    
    except PermissionError:
        return jsonify({"status": "error", "message": "Excel file is open. Please close it!"}), 500
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500